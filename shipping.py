# -*- coding: utf-8 -*-
"""
shipping.py — packaging, Excel export, email, and GitHub push.

- Builds XLSX from combined CSV with clickable links ("example_link" column).
  It tries, in order:
    1) PUBLIC_BASE_URL + a relative path computed from image_path (if possible)
    2) file:// local path (works if the file is on the viewer's machine or shared drive)
- Zips last 7 days of banner_screenshots.
- Sends weekly email (Monday only) with zip + ledger + xlsx attached.
- Commits and pushes to Git after each run if env configured.

Environment (set in Windows “User variables” or similar):
  GIT_REPO_DIR            e.g., C:\Users\tuguldur.kh\Downloads\adscraper-full-code
  GIT_REMOTE_NAME         default 'origin'
  GIT_BRANCH              default 'main'

  SMTP_HOST               e.g., smtp.gmail.com
  SMTP_PORT               e.g., 587
  SMTP_USER               your_smtp_username
  SMTP_PASS               your_app_password
  MAIL_FROM               "Ad Bot <bot@example.com>"
  MAIL_TO                 comma-separated list (e.g., "a@b.com,c@d.com")

  PUBLIC_BASE_URL         e.g., https://github.com/<USER>/<REPO>/blob/main
  OUTPUT_ROOT             (optional) override root for building relative file paths
"""

import os, csv, zipfile, smtplib, mimetypes, traceback, subprocess
from datetime import datetime, timedelta
from email.message import EmailMessage
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def _guess_output_root() -> str:
    # Prefer explicit env; else infer from typical structure
    env_root = os.getenv("OUTPUT_ROOT", "").strip()
    if env_root:
        return os.path.abspath(env_root)
    # Fallback: assume folder next to run.py called banner_screenshots
    here = Path(__file__).resolve().parent
    candidate = here / "banner_screenshots"
    return str(candidate)

def _to_rel(path_str: str, output_root: str) -> str:
    try:
        rel = os.path.relpath(path_str, output_root)
        return rel.replace("\\", "/")
    except Exception:
        return ""

def _public_url_from_rel(rel_path: str) -> str:
    base = (os.getenv("PUBLIC_BASE_URL") or "").rstrip("/")
    if not base or not rel_path:
        return ""
    return base + "/" + rel_path.lstrip("/")

def _file_url(local_path: str) -> str:
    # Excel will open file:// links when accessible
    p = Path(local_path).resolve()
    return "file:///" + str(p).replace("\\", "/")

def build_xlsx_from_csv(csv_path: str, xlsx_path: str) -> None:
    """
    Convert the combined CSV into XLSX with a clickable "example_link" column.
    We DO NOT modify your scrapers; we compute relative/URL on the fly.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "banners"

    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)

    if not os.path.exists(csv_path):
        wb.save(xlsx_path)
        print("[XLSX] CSV not found; created empty workbook.")
        return

    output_root = _guess_output_root()

    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []

        out_fields = list(fieldnames)
        if "example_link" not in out_fields:
            out_fields.append("example_link")

        ws.append(out_fields)

        for row in reader:
            # pick a local path column to build a link from (varies by scraper)
            path_candidate = row.get("example_path") or row.get("image_path") or ""
            link = ""

            if path_candidate:
                rel = _to_rel(path_candidate, output_root)
                url = _public_url_from_rel(rel)
                if url:
                    link = url
                else:
                    # fallback to file:// local link
                    link = _file_url(path_candidate)

            values = [row.get(k, "") for k in fieldnames]
            values.append(link)
            ws.append(values)

        # widths + hyperlink style
        for i, col in enumerate(out_fields, 1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(len(col) + 2, 18), 60)

        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=len(out_fields))
            url = str(cell.value or "")
            if url:
                cell.hyperlink = url
                cell.style = "Hyperlink"

    wb.save(xlsx_path)
    print(f"[XLSX] Wrote {xlsx_path}")

def zip_last_7_days(root_screenshots: str, out_zip_path: str) -> None:
    cutoff = datetime.now().date() - timedelta(days=7)
    os.makedirs(os.path.dirname(out_zip_path), exist_ok=True)
    with zipfile.ZipFile(out_zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for site in ("gogo.mn", "ikon.mn", "news.mn"):
            site_root = os.path.join(root_screenshots, site)
            if not os.path.isdir(site_root): 
                continue
            for day in os.listdir(site_root):
                # day folder like YYYY-MM-DD
                try:
                    d = datetime.strptime(day, "%Y-%m-%d").date()
                except Exception:
                    continue
                if d >= cutoff:
                    folder = os.path.join(site_root, day)
                    for dirpath, _, filenames in os.walk(folder):
                        for fn in filenames:
                            full = os.path.join(dirpath, fn)
                            rel  = os.path.relpath(full, root_screenshots)
                            zf.write(full, rel)
    print(f"[ZIP] Wrote {out_zip_path}")

def _attach_file(msg: EmailMessage, file_path: str):
    ctype, encoding = mimetypes.guess_type(file_path)
    if ctype is None or encoding is not None:
        ctype = "application/octet-stream"
    maintype, subtype = ctype.split("/", 1)
    with open(file_path, "rb") as f:
        msg.add_attachment(f.read(), maintype=maintype, subtype=subtype, filename=os.path.basename(file_path))

def send_email(subject: str, body: str, attachments: list[str]) -> None:
    smtp_host = os.getenv("SMTP_HOST", "")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_pass = os.getenv("SMTP_PASS", "")
    mail_from = os.getenv("MAIL_FROM", smtp_user)
    mail_to   = [x.strip() for x in os.getenv("MAIL_TO", "").split(",") if x.strip()]

    if not (smtp_host and smtp_user and smtp_pass and mail_to):
        print("[WARN] Email not sent (SMTP env missing)")
        return

    msg = EmailMessage()
    msg["From"] = mail_from
    msg["To"] = ", ".join(mail_to)
    msg["Subject"] = subject
    msg.set_content(body)

    for p in attachments:
        try:
            if os.path.exists(p):
                _attach_file(msg, p)
        except Exception:
            traceback.print_exc()

    with smtplib.SMTP(smtp_host, smtp_port) as s:
        s.starttls()
        s.login(smtp_user, smtp_pass)
        s.send_message(msg)
    print("[MAIL] Sent email to:", msg["To"])

def git_commit_and_push(repo_dir: str, message: str) -> None:
    """
    Uses local git in PATH. Repo should already have remote + auth set.
    Env (optional): GIT_REMOTE_NAME, GIT_BRANCH
    """
    remote = os.getenv("GIT_REMOTE_NAME", "origin")
    branch = os.getenv("GIT_BRANCH", "main")
    def run(*cmd):
        subprocess.run(cmd, cwd=repo_dir, check=False)
    run("git", "add", "-A")
    run("git", "commit", "-m", message)
    run("git", "push", remote, branch)
    print("[GIT] Push attempted to", remote, branch)
